from flask import Blueprint, render_template, make_response
from app.models import AllocationLog, Allocation, User, Task, Workload, EmployeeSkill
from app.decorators import login_required
from datetime import datetime
import io

log = Blueprint('log', __name__)


@log.route('/history')
@login_required
def history():
    logs = AllocationLog.query.order_by(AllocationLog.created_at.desc()).all()
    return render_template('log/history.html', logs=logs)


# ── PDF: Task Allocation Report ──────────────────────────────────────────────
@log.route('/download/task_allocation/pdf')
@login_required
def download_task_allocation_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    story   = []

    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=16, textColor=colors.HexColor('#1a2238'), spaceAfter=6)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey, spaceAfter=12)

    story.append(Paragraph("Task Allocation Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    allocations = Allocation.query.order_by(Allocation.allocated_at.desc()).all()

    data = [['#', 'Task Title', 'Assigned To', 'Priority', 'Skill Score', 'Status', 'Allocated On']]
    for i, a in enumerate(allocations, 1):
        data.append([
            str(i),
            a.task.title if a.task else '—',
            a.employee.name if a.employee else '—',
            (a.task.priority or '—').capitalize() if a.task else '—',
            str(a.skill_match_score or 0),
            (a.task.status or '—').capitalize() if a.task else '—',
            a.allocated_at.strftime('%d-%m-%Y') if a.allocated_at else '—'
        ])

    col_widths = [1*cm, 6*cm, 4*cm, 2.5*cm, 3*cm, 3*cm, 3.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#1a2238')),
        ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
        ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0,0), (-1,0), 9),
        ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE',     (0,1), (-1,-1), 8),
        ('GRID',         (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',   (0,0), (-1,-1), 6),
        ('BOTTOMPADDING',(0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=task_allocation_report.pdf'
    return response


# ── Excel: Task Allocation Report ────────────────────────────────────────────
@log.route('/download/task_allocation/excel')
@login_required
def download_task_allocation_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Task Allocation'

    header_fill = PatternFill('solid', fgColor='1a2238')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    border      = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    headers = ['#', 'Task Title', 'Assigned To', 'Priority', 'Skill Score', 'Status', 'Allocated On']
    col_w   = [5, 30, 20, 12, 14, 14, 18]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    allocations = Allocation.query.order_by(Allocation.allocated_at.desc()).all()
    for i, a in enumerate(allocations, 1):
        row_data = [
            i,
            a.task.title if a.task else '—',
            a.employee.name if a.employee else '—',
            (a.task.priority or '—').capitalize() if a.task else '—',
            a.skill_match_score or 0,
            (a.task.status or '—').capitalize() if a.task else '—',
            a.allocated_at.strftime('%d-%m-%Y') if a.allocated_at else '—'
        ]
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8F9FA'
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=j, value=val)
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=task_allocation_report.xlsx'
    return response


# ── PDF: Employee Performance Report ─────────────────────────────────────────
@log.route('/download/employee_performance/pdf')
@login_required
def download_employee_performance_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles      = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=16, textColor=colors.HexColor('#1a2238'), spaceAfter=6)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey, spaceAfter=12)

    story = []
    story.append(Paragraph("Employee Performance Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y, %H:%M')}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    employees   = User.query.filter_by(role='employee').all()
    allocations = Allocation.query.all()

    data = [['#', 'Employee', 'Email', 'Assigned Tasks', 'Completed', 'Completion %', 'Avg Skill Score', 'Current Workload']]
    for i, emp in enumerate(employees, 1):
        emp_allocs    = [a for a in allocations if a.employee_id == emp.id]
        assigned      = len(emp_allocs)
        completed     = sum(1 for a in emp_allocs if a.task and a.task.status == 'completed')
        completion_pct= f"{int((completed/assigned)*100)}%" if assigned > 0 else "0%"
        avg_score     = round(sum(a.skill_match_score for a in emp_allocs)/assigned, 1) if assigned > 0 else 0
        wl            = Workload.query.filter_by(employee_id=emp.id).first()
        wl_text       = f"{wl.current_tasks}/{wl.max_threshold}" if wl else "0/5"

        data.append([str(i), emp.name, emp.email, str(assigned),
                     str(completed), completion_pct, str(avg_score), wl_text])

    col_widths = [1*cm, 4*cm, 5.5*cm, 3.5*cm, 3*cm, 3.5*cm, 4*cm, 4*cm]
    table      = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1a2238')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 9),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=employee_performance_report.pdf'
    return response


# ── Excel: Employee Performance Report ───────────────────────────────────────
@log.route('/download/employee_performance/excel')
@login_required
def download_employee_performance_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Performance'

    header_fill = PatternFill('solid', fgColor='1a2238')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    border      = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    headers = ['#', 'Employee', 'Email', 'Assigned', 'Completed', 'Completion %', 'Avg Score', 'Workload']
    col_w   = [5, 22, 28, 12, 12, 15, 14, 14]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    employees   = User.query.filter_by(role='employee').all()
    allocations = Allocation.query.all()
    for i, emp in enumerate(employees, 1):
        emp_allocs     = [a for a in allocations if a.employee_id == emp.id]
        assigned       = len(emp_allocs)
        completed      = sum(1 for a in emp_allocs if a.task and a.task.status == 'completed')
        completion_pct = f"{int((completed/assigned)*100)}%" if assigned > 0 else "0%"
        avg_score      = round(sum(a.skill_match_score for a in emp_allocs)/assigned, 1) if assigned > 0 else 0
        wl             = Workload.query.filter_by(employee_id=emp.id).first()
        wl_text        = f"{wl.current_tasks}/{wl.max_threshold}" if wl else "0/5"

        row_data   = [i, emp.name, emp.email, assigned, completed, completion_pct, avg_score, wl_text]
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8F9FA'
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=j, value=val)
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=employee_performance_report.xlsx'
    return response


# ── PDF: Per Employee Task Report ────────────────────────────────────────────
@log.route('/download/employee/<int:emp_id>/pdf')
@login_required
def download_per_employee_pdf(emp_id):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    emp        = User.query.get_or_404(emp_id)
    emp_allocs = Allocation.query.filter_by(employee_id=emp_id).order_by(Allocation.allocated_at.desc()).all()

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                               leftMargin=1.5*cm, rightMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles      = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 fontSize=16, textColor=colors.HexColor('#1a2238'), spaceAfter=4)
    sub_style   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.grey, spaceAfter=12)

    story = []
    story.append(Paragraph(f"Task Report — {emp.name}", title_style))
    story.append(Paragraph(f"Email: {emp.email}  |  Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", sub_style))
    story.append(Spacer(1, 0.3*cm))

    data = [['#', 'Task Title', 'Priority', 'Skill Score', 'Status', 'Deadline', 'Allocated On']]
    for i, a in enumerate(emp_allocs, 1):
        data.append([
            str(i),
            a.task.title if a.task else '—',
            (a.task.priority or '—').capitalize() if a.task else '—',
            str(a.skill_match_score or 0),
            (a.task.status or '—').capitalize() if a.task else '—',
            str(a.task.deadline) if a.task and a.task.deadline else '—',
            a.allocated_at.strftime('%d-%m-%Y') if a.allocated_at else '—'
        ])

    col_widths = [1*cm, 7*cm, 3*cm, 3.5*cm, 3.5*cm, 4*cm, 4*cm]
    table      = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#1a2238')),
        ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,0), 9),
        ('ALIGN',         (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('FONTSIZE',      (0,1), (-1,-1), 8),
        ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
        ('TOPPADDING',    (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(table)
    doc.build(story)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={emp.name}_task_report.pdf'
    return response


# ── Excel: Per Employee Task Report ──────────────────────────────────────────
@log.route('/download/employee/<int:emp_id>/excel')
@login_required
def download_per_employee_excel(emp_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    emp        = User.query.get_or_404(emp_id)
    emp_allocs = Allocation.query.filter_by(employee_id=emp_id).order_by(Allocation.allocated_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = emp.name[:30]

    header_fill = PatternFill('solid', fgColor='1a2238')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    border      = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6')
    )

    headers = ['#', 'Task Title', 'Priority', 'Skill Score', 'Status', 'Deadline', 'Allocated On']
    col_w   = [5, 30, 14, 14, 14, 18, 18]
    for i, (h, w) in enumerate(zip(headers, col_w), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    for i, a in enumerate(emp_allocs, 1):
        row_data   = [
            i,
            a.task.title if a.task else '—',
            (a.task.priority or '—').capitalize() if a.task else '—',
            a.skill_match_score or 0,
            (a.task.status or '—').capitalize() if a.task else '—',
            str(a.task.deadline) if a.task and a.task.deadline else '—',
            a.allocated_at.strftime('%d-%m-%Y') if a.allocated_at else '—'
        ]
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8F9FA'
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i+1, column=j, value=val)
            cell.fill      = PatternFill('solid', fgColor=fill_color)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = border

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={emp.name}_task_report.xlsx'
    return response